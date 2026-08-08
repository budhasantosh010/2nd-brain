"""XLSX parser with sheet/range locators."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from second_brain.models import ParsedSegment
from second_brain.parsers.base import BaseParser


class XLSXParser(BaseParser):
    extensions = frozenset({".xlsx"})

    def parse(self, path: Path, source_id: str):  # type: ignore[no-untyped-def]
        workbook = load_workbook(path, read_only=True, data_only=True)
        segments: list[ParsedSegment] = []
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                chunk_size = 100
                max_col = max((len(row) for row in rows), default=1)
                for start in range(0, len(rows), chunk_size):
                    chunk = rows[start : start + chunk_size]
                    lines = ["\t".join("" if cell is None else str(cell) for cell in row) for row in chunk]
                    end_row = start + len(chunk)
                    locator = f"{sheet.title}!A{start + 1}:{get_column_letter(max_col)}{max(end_row, start + 1)}"
                    segments.append(
                        ParsedSegment(
                            segment_id=f"{source_id}:seg:{len(segments)}",
                            text="\n".join(lines),
                            locator=locator,
                            position=len(segments),
                            metadata={"sheet": sheet.title},
                        )
                    )
                if not rows:
                    segments.append(
                        ParsedSegment(
                            segment_id=f"{source_id}:seg:{len(segments)}",
                            text="",
                            locator=f"{sheet.title}!A1:A1",
                            position=len(segments),
                            metadata={"sheet": sheet.title},
                        )
                    )
        finally:
            workbook.close()
        return self.document(
            path,
            source_id,
            segments,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
